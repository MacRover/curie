import serial  # pip install pyserial
import time
import keyboard  # pip install keyboard
import matplotlib.pyplot as plt # pip install matplotlib
import datetime
import csv
from openpyxl import Workbook, load_workbook

"""
Where press is specified, press and hold


Press and hold t to see temperature graph
Press and hold h to see humidity graph
Press and hold p to see pH graph
press and hold e to see ec graph

Press and hold o to change an offset
Press and hold s to change a scale factor

Press and hold y to see temperature moving average graph
Press and hold j to see humidity moving average graph
Press and hold q to see pH moving average graph
Press and hold r to see ec moving average graph


"""


"""

The COM port of the CH340 Transceiver MUST be updated here


"""

PORT = "COM5"  # Change this to your COM port

"""

The number of moving average points can be changed here

"""
MOVING_AVERAGE_POINTS = 20


"""
Change Offset and Scaling parameters here
"""

TEMP_OFFSET_BASE = 0
HUMIDITY_OFFSET_BASE = 0
PH_OFFSET_BASE = 0
EC_OFFSET_BASE = 0

TEMP_SCALING_BASE = 1
HUMIDITY_SCALING_BASE = 1
PH_SCALING_BASE = 1
EC_SCALING_BASE = 1



temp_offset = TEMP_OFFSET_BASE
ph_offset = PH_OFFSET_BASE
humidity_offset = HUMIDITY_OFFSET_BASE
ec_offset = EC_OFFSET_BASE

temp_scaling = TEMP_SCALING_BASE
ph_scaling = PH_SCALING_BASE
humidity_scaling = HUMIDITY_SCALING_BASE
ec_scaling = EC_SCALING_BASE

def change_offset(temp_offset, humidity_offset, ph_offset, ec_offset):

    print("Enter a command")
    print("1: Change temperature offset")
    print("2: Change humidity offset")
    print("3: Change pH offset")
    print("4: Change electrical conductivity offset")
        

    cmd = int(input(""))


    if (cmd == 1):
        temp_offset = float(input("enter value: "))

    elif (cmd == 2):
        humidity_offset = float(input("enter value: "))

    elif (cmd == 3):
        ph_offset = float(input("enter value: "))

    elif (cmd == 4):
        ec_offset = float(input("enter value: "))

    return temp_offset, humidity_offset, ph_offset, ec_offset
    



def change_scaling(temp_scaling, humidity_scaling, ph_scaling, ec_scaling):

    print("Enter a command")
    print("1: Change temperature scaling")
    print("2: Change humidity scaling")
    print("3: Change pH scaling")
    print("4: Change electrical conductivity scaling")
    

    cmd = int(input(""))

    if (cmd == 1):
        temp_scaling = float(input("enter value: "))

    elif (cmd == 2):
        humidity_scaling = float(input("enter value: "))

    elif (cmd == 3):
        ph_scaling = float(input("enter value: "))

    elif (cmd == 4):
        ec_scaling = float(input("enter value: "))

    return temp_scaling, humidity_scaling, ph_scaling, ec_scaling


def offset_and_scale(scale, offset, value):

    return scale*value + offset


def moving_average(array, window):

    # Calculates the moving average very straightforwardly

    if (len(array)) < window:
        return sum(array) / len(array)
    else:
        return sum(array[-window:]) / window # Return the last window points




# Functions from Arduino

def crc16_2(buf: bytes) -> int:
    crc = 0xFFFF
    for b in buf:
        crc ^= b
        for _ in range(8):
            if crc & 0x0001:
                crc >>= 1
                crc ^= 0xA001
            else:
                crc >>= 1

    # byte swap (same as Arduino)
    crc = ((crc & 0x00FF) << 8) | ((crc & 0xFF00) >> 8)
    return crc

import time

def read_n(ser, length, timeout_ms=500) -> bytes:
    data = bytearray()
    start = time.time()

    while len(data) < length:
        if ser.in_waiting:
            data += ser.read(1)
        if (time.time() - start) * 1000 > timeout_ms:
            break

    return bytes(data)

def read_humiture_ecph(ser, command: bytes):
    """
    ser     : opened serial.Serial object
    command : 8-byte command frame (your Com array)
    """

    while True:
        time.sleep(0.1)

        # Send command
        ser.write(command)
        ser.flush()

        # Expect header 01 03 08
        ch = read_n(ser, 1)
        if len(ch) != 1 or ch[0] != 0x01:
            continue

        ch = read_n(ser, 1)
        if len(ch) != 1 or ch[0] != 0x03:
            continue

        ch = read_n(ser, 1)
        if len(ch) != 1 or ch[0] != 0x08:
            continue

        # Read remaining bytes
        rest = read_n(ser, 10)
        if len(rest) != 10:
            continue

        data = bytes([0x01, 0x03, 0x08]) + rest

        # CRC check
        crc_calc = crc16_2(data[:11])
        crc_recv = (data[11] << 8) | data[12]

        if crc_calc != crc_recv:
            continue

        # Parse values
        hum = (data[3] << 8 | data[4]) / 10.0
        tem = (data[5] << 8 | data[6]) / 10.0
        ec  = (data[7] << 8 | data[8])
        ph  = (data[9] << 8 | data[10]) / 10.0

        return hum, tem, ec, ph


# ---------------- UART Setup ----------------
BAUD = 9600
BYTESIZE = 8
PARITY = serial.PARITY_NONE
STOPBITS = 1
TIMEOUT = 0
ser = serial.Serial(port = PORT, baudrate = BAUD, bytesize = BYTESIZE, parity = PARITY, stopbits = STOPBITS, timeout = TIMEOUT)

SHEET_NAME_INSTANT = "Instantaneous Values"
SHEET_NAME_MOVING_AVERAGES = "Moving Averages"

# Original 8 byte command

Com = bytes([0x01, 0x03, 0x00, 0x00, 0x00, 0x04, 0x44, 0x09])
# ---------------- Data storage ----------------
temp = []
hum = []
ph = []
ec = []
x = []

# Storage for moving averages 
temp_avg = []
hum_avg = []
ph_avg = []
ec_avg = []

start = time.time()
last_plot = 0  # For key debounce

excel_file = "data_" + str(datetime.datetime.now()).replace(" ", "_").replace(":", ".") + ".xlsx"



# Workbook creation
wb = Workbook()

# Deleting the default sheet

del wb["Sheet"]

ws_inst = wb.create_sheet(SHEET_NAME_INSTANT)
ws_avg = wb.create_sheet(SHEET_NAME_MOVING_AVERAGES)

ws_inst.append(["Time (s)", "Temperature", "Humidity", "pH", "Electrical Conductivity"])
ws_avg.append(["Time (s)", "Temperature", "Humidity", "pH", "Electrical Conductivity"])

wb.save(excel_file)

prev_now = 0

print("Listening...")

plot_total_open_time = 0

try:
    while True:
        # ---------- Read UART ----------
        # line = ser.readline().decode("utf-8", errors="ignore").strip()
        # if line:
        #     try:
        #         values = [float(v) for v in line.split(" ")]
        #         if len(values) == 4:
        #             a, b, c, d = values
        #             temp.append(a)
        #             hum.append(b)
        #             ph.append(c)
        #             ec.append(d)
        #             now = time.time() - start
        #             x.append(now)
        #             print("Temperature: " + str(a) + " | Humidity: " + str(b) + " | pH: " + str(c) + " | Electrical Conductivity: " + str(d))

                    # with open(csv_file, mode="a", newline = "") as f:
                    #     writer = csv.writer(f)
                    #     writer.writerow([now, a, b, c, d])


        #     except ValueError:
        #         pass  # ignore malformed lines

        hum_val, temp_val, ec_val, ph_val = read_humiture_ecph(ser, Com)

        temp_val = offset_and_scale(temp_scaling, temp_offset, temp_val)
        hum_val = offset_and_scale(humidity_scaling, humidity_offset, hum_val)
        ph_val = offset_and_scale(ph_scaling, ph_offset, ph_val)
        ec_val = offset_and_scale(ec_scaling, ec_offset, ec_val)

        """

        Uncomment this line to view instantaneous values
        

        """
        # print("Temperature: " + str(temp_val) + " | Humidity: " + str(hum_val) + " | pH: " + str(ec_val) + " | Electrical Conductivity: " + str(ph_val))


        # Append it all to the lists

        temp.append(temp_val)
        hum.append(hum_val)
        ec.append(ec_val)
        ph.append(ph_val)

        avg_temp = moving_average(temp, MOVING_AVERAGE_POINTS)
        avg_hum = moving_average(hum, MOVING_AVERAGE_POINTS)
        avg_ph = moving_average(ph, MOVING_AVERAGE_POINTS)
        avg_ec = moving_average(ec, MOVING_AVERAGE_POINTS)

        temp_avg.append(avg_temp)
        hum_avg.append(avg_hum)
        ph_avg.append(avg_ph)
        ec_avg.append(avg_ec)

        """

        Uncomment this line to view moving averages
        
        """

        print("Temperature: " + str(avg_temp) + " | Humidity: " + str(avg_hum) + " | pH: " + str(avg_ph) + " | Electrical Conductivity: " + str(avg_ec))


        # Need to subtract plot_total_open_time to fix the plots being a complete mess since the timers don't stop
        now = time.time() - start - plot_total_open_time
        x.append(now)

        # csv_writing

        ws_inst.append([now, temp_val, hum_val, ph_val, ec_val])
        ws_avg.append([now, avg_temp, avg_hum, avg_ph, avg_ec])

        # Saving logic
        if ((now - prev_now) >= 30):
            wb.save(excel_file)
            prev_now = now
            print("Excel File Saved")
            opened_file = False


        # ---------- Check for 't' press ----------
        if keyboard.is_pressed("t") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, temp, color='red', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Temperature")
            plt.title("Temperature vs Time")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("h") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, hum, color='blue', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Humidity")
            plt.title("Humidity vs Time")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("p") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, ph, color='magenta', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("pH")
            plt.title("pH vs Time")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("e") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, ec, color='orange', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Electrical Conductivity")
            plt.title("Electrical Conductivity vs Time")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("y") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, temp_avg, color='red', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Temperature")
            plt.title("Temperature vs Time (M.AVG)")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("j") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, hum_avg, color='blue', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Humidity")
            plt.title("Humidity vs Time (M.AVG)")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("q") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, ph_avg, color='magenta', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("pH")
            plt.title("pH vs Time (M.AVG)")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot

        if keyboard.is_pressed("r") and time.time() - last_plot > 0.5:
            last_plot = time.time()
            # Pause main loop and show snapshot plot
            plt.figure()
            plt.plot(x, ec_avg, color='orange', lw=2)
            plt.xlabel("Time (s)")
            plt.ylabel("Electrical Conductivity")
            plt.title("Electrical Conductivity vs Time (M.AVG)")
            plt.grid(True)
            plt.show()  # Blocks here until window closed
            print("Plot closed, resuming data collection...")
            plot_close_time = time.time()
            plot_total_open_time += plot_close_time - last_plot


        if keyboard.is_pressed("o") and time.time() - last_plot > 0.5:
            offset_select_time = time.time()
            temp_offset, humidity_offset, ph_offset, ec_offset = change_offset(temp_offset, humidity_offset, ph_offset, ec_offset)
            offset_close_time = time.time()
            plot_total_open_time += offset_close_time - offset_select_time

        if keyboard.is_pressed("s") and time.time() - last_plot > 0.5:
            offset_select_time = time.time()
            temp_scaling, humidity_scaling, ph_scaling, ec_scaling = change_scaling(temp_scaling, humidity_scaling, ph_scaling, ec_scaling)
            offset_close_time = time.time()
            plot_total_open_time += offset_close_time - offset_select_time

except KeyboardInterrupt:
    print("Exiting...")
    wb.save(excel_file)
    ser.close()



